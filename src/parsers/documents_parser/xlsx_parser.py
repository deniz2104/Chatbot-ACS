import logging

import openpyxl
from langchain_core.documents import Document

from src.parsers.documents_parser.spreadsheet_shared import sheet_blocks_to_docs
from src.parsers.entries import DocumentEntry
from src.parsers.error_handlers import parse_error

logger = logging.getLogger(__name__)

_LABEL = "XLSX"

def _sheet_to_text(ws) -> str:
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(cell).strip() if cell is not None else "" for cell in row]
        rows.append("\t".join(cells))
    return "\n".join(rows)

def _is_meaningful_sheet(ws) -> bool:
    return any(
        cell.value not in ("", None)
        for row in ws.iter_rows()
        for cell in row
    )

def process_xlsx(document_entry: DocumentEntry) -> list[Document]:
    path = document_entry.local_path
    logger.info("[XLSX] Loading file: %s", path)

    wb = None
    with parse_error(_LABEL, path):
        wb = openpyxl.load_workbook(path, data_only=True)
    if wb is None:
        return []

    sheet_blocks = []
    for ws in wb.worksheets:
        if not _is_meaningful_sheet(ws):
            logger.debug("[XLSX] Skipping empty sheet: %s", ws.title)
            continue
        sheet_blocks.append(f"Sheet: {ws.title}\n{_sheet_to_text(ws)}")

    wb.close()

    if not sheet_blocks:
        logger.warning("[XLSX] No meaningful sheets found in %s", path)
        return []

    return sheet_blocks_to_docs(sheet_blocks, document_entry, path, _LABEL)